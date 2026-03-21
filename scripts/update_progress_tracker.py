#!/usr/bin/env python3
from __future__ import annotations

import argparse
import subprocess
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = REPO_ROOT / "output" / "spreadsheet" / "startpray-progress-tracker.xlsx"


def run_git(args: list[str]) -> str:
    try:
        result = subprocess.run(
            ["git", *args],
            cwd=REPO_ROOT,
            check=True,
            capture_output=True,
            text=True,
        )
        return result.stdout.strip()
    except subprocess.CalledProcessError:
        return ""


def get_git_context() -> dict[str, str]:
    branch = run_git(["rev-parse", "--abbrev-ref", "HEAD"]) or "(unknown-branch)"
    commit = run_git(["rev-parse", "--short", "HEAD"]) or "(unknown-commit)"
    full_commit = run_git(["rev-parse", "HEAD"]) or commit
    latest_tag = run_git(["describe", "--tags", "--abbrev=0"]) or "(no tag)"
    message = run_git(["log", "-1", "--pretty=%s"]) or "(no message)"
    remote = run_git(["config", "--get", "remote.origin.url"]) or "(no remote)"
    return {
        "branch": branch,
        "commit": commit,
        "full_commit": full_commit,
        "latest_tag": latest_tag,
        "message": message,
        "remote": remote,
    }


def ensure_push_log_sheet(workbook):
    if "Push_Log" in workbook.sheetnames:
        return workbook["Push_Log"]

    ws = workbook.create_sheet("Push_Log")
    headers = [
        "Timestamp",
        "Event",
        "Branch",
        "Commit",
        "Tag",
        "Commit Message",
        "Remote",
        "Refs",
    ]
    ws.append(headers)
    widths = {
        "A": 22,
        "B": 12,
        "C": 14,
        "D": 12,
        "E": 16,
        "F": 48,
        "G": 36,
        "H": 44,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    return ws


def append_weekly_log(workbook, event: str, git_ctx: dict[str, str], refs: str):
    if "Weekly_Log" not in workbook.sheetnames:
        return
    ws = workbook["Weekly_Log"]
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    if event == "push":
        summary = f"[AUTO] Push recorded: {git_ctx['branch']} @ {git_ctx['commit']}"
        blockers = ""
        next_plan = "Review release status and deployment result."
    elif event == "manual":
        summary = f"[AUTO] Manual sync: {git_ctx['branch']} @ {git_ctx['commit']}"
        blockers = ""
        next_plan = "Update task status and release mapping."
    else:
        summary = f"[AUTO] Commit recorded: {git_ctx['branch']} @ {git_ctx['commit']}"
        blockers = ""
        next_plan = "Continue task updates and prepare push."
    completed_ids = ""
    owner = "git-hook"
    ws.append([now, summary, completed_ids, blockers, next_plan, owner])


def append_push_log(workbook, event: str, git_ctx: dict[str, str], refs: str):
    ws = ensure_push_log_sheet(workbook)
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.append(
        [
            now,
            event,
            git_ctx["branch"],
            git_ctx["commit"],
            git_ctx["latest_tag"],
            git_ctx["message"],
            git_ctx["remote"],
            refs or "",
        ]
    )


def update_dashboard(workbook, git_ctx: dict[str, str]):
    if "Dashboard" not in workbook.sheetnames:
        return
    ws = workbook["Dashboard"]
    ws["F3"] = "Current Branch"
    ws["G3"] = git_ctx["branch"]
    ws["F4"] = "Current Commit"
    ws["G4"] = git_ctx["commit"]
    ws["F5"] = "Latest Tag"
    ws["G5"] = git_ctx["latest_tag"]
    ws["F6"] = "Snapshot Date"
    ws["G6"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def update_tasks_git_version(workbook, git_ctx: dict[str, str]):
    if "Tasks" not in workbook.sheetnames:
        return
    ws = workbook["Tasks"]
    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    status_col = headers.get("Status")
    git_col = headers.get("Git Version")
    if not status_col or not git_col:
        return

    for row in range(2, ws.max_row + 1):
        status = str(ws.cell(row=row, column=status_col).value or "").strip()
        current = str(ws.cell(row=row, column=git_col).value or "").strip()
        if status in {"In Progress", "Done"} and (not current or current == "(unknown-commit)"):
            ws.cell(row=row, column=git_col).value = git_ctx["commit"]


def update_release_actual_git(workbook, event: str, git_ctx: dict[str, str]):
    if event != "push":
        return
    if "Release_Tracker" not in workbook.sheetnames:
        return
    ws = workbook["Release_Tracker"]
    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    status_col = headers.get("Status")
    actual_col = headers.get("Actual Git")
    if not status_col or not actual_col:
        return

    # Update the last active release row first; if none active, keep as log-only.
    for row in range(ws.max_row, 1, -1):
        status = str(ws.cell(row=row, column=status_col).value or "").strip()
        if status in {"In Progress", "Planned"}:
            ws.cell(row=row, column=actual_col).value = git_ctx["commit"]
            break


def main():
    parser = argparse.ArgumentParser(description="Update progress tracker workbook from git state.")
    parser.add_argument("--event", choices=["commit", "push", "manual"], default="manual")
    parser.add_argument("--refs", default="", help="Optional refs payload from git hook stdin.")
    args = parser.parse_args()

    if not WORKBOOK_PATH.exists():
        raise SystemExit(f"Workbook not found: {WORKBOOK_PATH}")

    git_ctx = get_git_context()
    workbook = load_workbook(WORKBOOK_PATH)

    update_dashboard(workbook, git_ctx)
    update_tasks_git_version(workbook, git_ctx)
    update_release_actual_git(workbook, args.event, git_ctx)
    append_weekly_log(workbook, args.event, git_ctx, args.refs)
    append_push_log(workbook, args.event, git_ctx, args.refs)

    workbook.save(WORKBOOK_PATH)
    print(f"Updated workbook: {WORKBOOK_PATH}")
    print(f"Event={args.event} Branch={git_ctx['branch']} Commit={git_ctx['commit']}")


if __name__ == "__main__":
    main()
